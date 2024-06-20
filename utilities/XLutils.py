import openpyxl
from openpyxl.styles import PatternFill


class XLutils:
    file_path = r"C:\Users\vamsh\PycharmProjects\openKartV1\testData\DataDriven.xlsx"
    def getRowCount(self, file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(self, file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)

    def readData(self,file, sheetName, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rownum, colnum).value

    def writeData(self, file , sheetName, rownum, colnum , data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rownum, colnum).value = data
        workbook.save(file)

    def fillGreeenColor(self, file, sheetName, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        greenFill = PatternFill(start_color='60b212',
                                end_color='60b212',
                                fill_type='solid')
        sheet.cell(rownum, colnum).fill = greenFill
        workbook.save(file)

    def fillRedColor(self, file, sheetName, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        redFill = PatternFill(start_color='ff0000',
                                end_color='ff0000',
                                fill_type='solid')
        sheet.cell(rownum, colnum).fill = redFill
        workbook.save(file)

excel = XLutils()
print(excel.getRowCount(XLutils.file_path,"Sheet1"))
print(excel.getColumnCount(XLutils.file_path, "Sheet1"))
print(excel.readData(XLutils.file_path, "Sheet1", 2 ,1))
print(excel.fillGreeenColor(XLutils.file_path, "Sheet1", 2, 3))
print(excel.fillRedColor(XLutils.file_path, "Sheet1", 2, 3))

